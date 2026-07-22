"""Matriz funil x funil: nossas oportunidades x os cards do CRM do HubSoft.

Nao existe API de cards do HubSoft (o /crm/all so lista boards, todo drill da
404), entao a unica fonte do lado deles e o export .xlsx da Gabi. Este service:

  1. parseia a planilha (parse_planilha)
  2. classifica a etapa deles em ganho/aberto/perdido (situacao_deles)
  3. cruza por id_prospecto contra as nossas OportunidadeVenda AO VIVO e monta a
     matriz de concordancia, os duplicados e o que so existe de um lado
     (montar_aba)

A planilha e um retrato do momento do envio; o NOSSO lado e recalculado a cada
visita, entao movimentacao no CRM aparece na hora e a matriz nunca fica velha
do nosso lado.

Chave de cruzamento: o `id_prospecto` do card contra o `id_hubsoft` do nosso
lead (que E o id do prospecto que criamos no HubSoft: foi assim que 24211 do
nosso lado casou com o card 24210). Custou uma sessao provar que ela e melhor
que CPF (casou 763 contra 232), porque a maioria dos cards nao tem CPF.
"""
from __future__ import annotations

import re
from collections import Counter

# Situacoes canonicas dos dois lados
GANHO, ABERTO, PERDIDO = 'ganho', 'aberto', 'perdido'

# Colunas que o parser aproveita da planilha. As demais (email, valor, plano,
# a data_cadastro_prospecto duplicada) sao ignoradas de proposito.
COLUNAS = (
    'id_prospecto', 'crm', 'crm_etapa', 'status_prospecto', 'nome_cartao',
    'nome_prospecto', 'prospecto_cpf_cnpj', 'prospecto_telefone', 'tag',
    'usuario', 'data_cadastro_cartao',
)

# Etapas do HubSoft que contam como perda. O resto que nao e "CADASTRO APROVADO"
# fica em aberto (negociacao viva): ASSUNTOS COMERCIAIS, LOJA VIRTUAL NUVYON,
# CAPTACAO DE CLIENTE, ANALISE DE VIABILIDADE.
_PERDIDO_EXATO = {'CREDITO NEGADO', 'VIABILIDADE NEGATIVA'}


def _digitos(v) -> str:
    return re.sub(r'\D', '', str(v or ''))


def _norm_id(v) -> str:
    """id_prospecto vem numero na planilha (24291 ou 24291.0) e texto no nosso
    lead. Normaliza os dois pro mesmo formato so digitos."""
    s = str(v or '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    return _digitos(s)


def situacao_deles(crm_etapa: str) -> str:
    """Traduz a etapa do CRM deles nas nossas tres situacoes."""
    e = (crm_etapa or '').strip().upper()
    if e == 'CADASTRO APROVADO':
        return GANHO
    if e.startswith('DESIST') or e in _PERDIDO_EXATO:
        return PERDIDO
    return ABERTO


def _situacao_nossa(estagio) -> str:
    if estagio and estagio.is_final_ganho:
        return GANHO
    if estagio and estagio.is_final_perdido:
        return PERDIDO
    return ABERTO


def parse_planilha(arquivo) -> list[dict]:
    """Le o .xlsx do CRM e devolve um card normalizado por linha.

    Tolerante a coluna faltando (outro export pode nao ter `tag`/`usuario`) e a
    header repetido (a planilha traz `data_cadastro_prospecto` duas vezes): usa
    a primeira ocorrencia de cada nome. Descarta linha sem id_prospecto, que nao
    da pra casar.
    """
    import openpyxl

    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = ws.iter_rows(values_only=True)
    try:
        cabecalho = [str(h).strip() if h is not None else '' for h in next(linhas)]
    except StopIteration:
        return []

    # primeira ocorrencia de cada coluna que interessa
    idx = {}
    for col in COLUNAS:
        if col in cabecalho:
            idx[col] = cabecalho.index(col)

    def val(linha, col):
        i = idx.get(col)
        if i is None or i >= len(linha):
            return ''
        return '' if linha[i] is None else str(linha[i]).strip()

    cards = []
    for linha in linhas:
        idp = _norm_id(val(linha, 'id_prospecto'))
        if not idp:
            continue
        etapa = val(linha, 'crm_etapa')
        cards.append({
            'id_prospecto': idp,
            'crm': val(linha, 'crm'),
            'crm_etapa': etapa,
            'situacao': situacao_deles(etapa),
            'status_prospecto': val(linha, 'status_prospecto'),
            'nome_cartao': val(linha, 'nome_cartao'),
            'nome_prospecto': val(linha, 'nome_prospecto'),
            'cpf': _digitos(val(linha, 'prospecto_cpf_cnpj')),
            'telefone': _digitos(val(linha, 'prospecto_telefone')),
            'tag': val(linha, 'tag'),
            'usuario': val(linha, 'usuario'),
            'data_cadastro_cartao': val(linha, 'data_cadastro_cartao'),
        })
    wb.close()
    return cards


def _importacao_atual(tenant):
    from apps.integracoes.models import ImportacaoCRMHubsoft
    return (ImportacaoCRMHubsoft.all_tenants
            .filter(tenant=tenant).order_by('-enviado_em').first())


def _indices_nossos(tenant):
    """Nossos indices, lidos uma vez: oportunidade por id_hubsoft do lead
    (situacao ao vivo), e lead por cpf/telefone (pra achar duplicata) com o
    id_hubsoft. O id_hubsoft do lead E o id do prospecto que criamos, entao ele
    casa direto com o id_prospecto do card."""
    from apps.comercial.leads.models import LeadProspecto
    from apps.comercial.crm.models import OportunidadeVenda

    op_por_prosp = {}
    for o in (OportunidadeVenda.all_tenants
              .filter(tenant=tenant)
              .select_related('estagio', 'lead')):
        idp = _norm_id(o.lead.id_hubsoft) if o.lead_id else ''
        if idp:
            op_por_prosp[idp] = {
                'situacao': _situacao_nossa(o.estagio),
                'lead_id': o.lead_id,
                'estagio': o.estagio.nome if o.estagio else '',
                'nome': (o.lead.nome_razaosocial or '') if o.lead_id else '',
            }

    lead_por_cpf, lead_por_tel = {}, {}
    for l in LeadProspecto.all_tenants.filter(tenant=tenant):
        info = {
            'lead_id': l.pk,
            'nome': l.nome_razaosocial or '',
            'id_hubsoft': _norm_id(l.id_hubsoft),
            'status_api': l.status_api,
        }
        cpf = _digitos(l.cpf_cnpj)
        if cpf:
            lead_por_cpf.setdefault(cpf, info)
        tel = _digitos(l.telefone)
        if len(tel) >= 10:
            lead_por_tel.setdefault(tel[-8:], info)
    return op_por_prosp, lead_por_cpf, lead_por_tel


def montar_aba(tenant) -> dict:
    """Tudo que a aba Oportunidades precisa. Sem cache: e leitura local (DB +
    JSON), rapida, e o valor e justamente refletir o CRM ao vivo."""
    imp = _importacao_atual(tenant)
    if imp is None:
        return {'tem_import': False}

    todos = imp.cards or []

    # id_prospecto = 0 e o card aberto direto no CRM sem nunca linkar um prospecto
    # no HubSoft (cpf "Nao Possui prospecto", sem telefone). Nao tem chave nenhuma
    # pra casar, entao vira bucket proprio em vez de inchar "so existem la" e a
    # cobertura. Na planilha de 21/07 eram 382 dos 1226.
    sem_prospecto = [c for c in todos if c['id_prospecto'] in ('', '0')]
    cards = [c for c in todos if c['id_prospecto'] not in ('', '0')]

    op_por_prosp, lead_por_cpf, lead_por_tel = _indices_nossos(tenant)

    matriz = Counter()          # (situacao_deles, situacao_nossa) -> n
    divergentes, duplicados, so_deles = [], [], []
    casados = 0

    for c in cards:
        nosso = op_por_prosp.get(c['id_prospecto'])
        if nosso:
            casados += 1
            matriz[(c['situacao'], nosso['situacao'])] += 1
            if c['situacao'] != nosso['situacao']:
                divergentes.append({**c,
                                    'nossa_situacao': nosso['situacao'],
                                    'nosso_estagio': nosso['estagio'],
                                    'lead_id': nosso['lead_id']})
            continue

        # nao casou por id_prospecto: a pessoa existe aqui com OUTRO id (duplicata)?
        lead = (lead_por_cpf.get(c['cpf']) if c['cpf'] else None)
        if lead is None and len(c['telefone']) >= 10:
            lead = lead_por_tel.get(c['telefone'][-8:])
        if lead and lead['id_hubsoft'] and lead['id_hubsoft'] != c['id_prospecto']:
            duplicados.append({**c,
                               'lead_id': lead['lead_id'],
                               'nosso_id_hubsoft': lead['id_hubsoft'],
                               'nosso_nome': lead['nome']})
        else:
            so_deles.append(c)

    # cobertura sobre os cards REAIS (com prospecto), nao sobre os 1226 crus, pra
    # os 382 sem_prospecto nao derrubarem o percentual
    total = len(cards)
    so_nossos = sum(1 for idp in op_por_prosp
                    if idp not in {c['id_prospecto'] for c in cards})

    # concordancia = diagonal da matriz
    concordam = sum(n for (d, nn), n in matriz.items() if d == nn)

    # Uma lista unica de problemas, com a categoria por linha, pra a tela usar UMA
    # datatable com chips (mesmo padrao da aba Vendas) em vez de 4 tabelas.
    problemas = _unificar(divergentes, duplicados, so_deles, sem_prospecto)
    chips = [
        {'slug': 'Divergência', 'total': len(divergentes)},
        {'slug': 'Duplicado', 'total': len(duplicados)},
        {'slug': 'Só existe lá', 'total': len(so_deles)},
        {'slug': 'Sem prospecto', 'total': len(sem_prospecto)},
    ]

    return {
        'tem_import': True,
        'enviado_em': imp.enviado_em,
        'nome_arquivo': imp.nome_arquivo,
        'total_cards': len(todos),
        'total': total,
        'casados': casados,
        'cobertura_pct': round(100.0 * casados / total) if total else 0,
        'total_so_deles': len(so_deles),
        'total_duplicados': len(duplicados),
        'total_sem_prospecto': len(sem_prospecto),
        'so_nossos': so_nossos,
        'concordam': concordam,
        'concordancia_pct': round(100.0 * concordam / casados) if casados else 0,
        'total_divergentes': len(divergentes),
        'matriz': _matriz_para_template(matriz),
        'problemas': problemas,
        'total_problemas': len(problemas),
        'chips': [c for c in chips if c['total']],
    }


def _linha(card, categoria, detalhe='', lead_id=None, tag=None):
    return {
        'categoria': categoria,
        'nome': card.get('nome_cartao') or card.get('nome_prospecto') or '',
        'id_prospecto': card['id_prospecto'],
        'crm_etapa': card['crm_etapa'],
        'detalhe': detalhe,
        'tag': card.get('tag', '') if tag is None else tag,
        'usuario': card.get('usuario', ''),
        'lead_id': lead_id,
    }


def _unificar(divergentes, duplicados, so_deles, sem_prospecto) -> list:
    """Achata os quatro baldes numa lista so, cada linha com sua categoria e um
    campo `detalhe` que carrega o que e especifico de cada tipo (nosso estagio,
    nosso lead, telefone ou equipe)."""
    linhas = []
    for d in divergentes:
        linhas.append(_linha(d, 'Divergência', detalhe=f"Nosso: {d['nosso_estagio']}",
                             lead_id=d.get('lead_id')))
    for x in duplicados:
        linhas.append(_linha(x, 'Duplicado',
                             detalhe=f"Nosso lead #{x['lead_id']} (id_hs {x['nosso_id_hubsoft']})",
                             lead_id=x.get('lead_id')))
    for x in so_deles:
        linhas.append(_linha(x, 'Só existe lá',
                             detalhe=f"tel {x['telefone']}" if x.get('telefone') else ''))
    for x in sem_prospecto:
        # sem prospecto nao tem tag util; a equipe (board) e o que importa
        linhas.append(_linha(x, 'Sem prospecto', detalhe=x.get('crm', ''), tag=x.get('crm', '')))
    return linhas


_ROTULO = {GANHO: 'Ganho', ABERTO: 'Aberto', PERDIDO: 'Perdido'}


def _matriz_para_template(matriz: Counter) -> dict:
    """Serializa a matriz 3x3 num formato que o template le sem logica: cada
    celula ja vem com o que destacar (diagonal = concordancia; deles ABERTO /
    nos PERDIDO = perda prematura, a divergencia que importa)."""
    ordem = (GANHO, ABERTO, PERDIDO)
    linhas = []
    for d in ordem:
        celulas = []
        for n in ordem:
            v = matriz.get((d, n), 0)
            celulas.append({
                'valor': v,
                'concorda': d == n,
                'perda_prematura': d == ABERTO and n == PERDIDO and v > 0,
            })
        linhas.append({
            'deles': d, 'deles_label': _ROTULO[d],
            'celulas': celulas, 'total': sum(c['valor'] for c in celulas),
        })
    col_total = [sum(matriz.get((d, n), 0) for d in ordem) for n in ordem]
    return {
        'colunas': [_ROTULO[n] for n in ordem],
        'linhas': linhas,
        'col_total': col_total,
        'total': sum(matriz.values()),
        'perda_prematura': matriz.get((ABERTO, PERDIDO), 0),
    }
